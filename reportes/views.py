from django.shortcuts import render
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import colors, Font
from django.http.response import HttpResponse
from ausencias.models import Ausencia
from funcionarios.models import Funcionario
import datetime
from relativedeltafield import relativedelta
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required


# Create your views here.
@method_decorator(login_required(), name='dispatch')
class ReporteList(TemplateView):
    template_name = "reportes/reportes.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, {'title': 'REPORTES'})

# REPORTES LICENCIAS ################


# REPORTE EXCEL LICENCIAS CORRESPONDIENTES AL MES ACTUAL
@method_decorator(login_required(), name='dispatch')
class ReporteLicenciasExcelMesActual(TemplateView):
    def get(self, request, *args, **kwargs):
        hoy = datetime.datetime.now()
        annio = hoy.year
        mes = hoy.month
        ausencias = Ausencia.objects.filter(fecha_inicio__year=annio).filter(fecha_inicio__month=mes)
        wb = Workbook()
        ws = wb.active
        # TRABAJANDO EN EL REPORTE EXCEL
        b1 = ws['B1']
        b3 = ws['B3']
        c3 = ws['C3']
        d3 = ws['D3']
        e3 = ws['E3']
        f3 = ws['F3']
        ftr = Font(color=colors.BLUE, bold=True)
        ftb = Font(color=colors.BLACK, bold=True, size=10)
        b1.font = ftr
        b3.font = ftb
        c3.font = ftb
        d3.font = ftb
        e3.font = ftb

        ws['B1'] = 'REPORTE DE LICENCIAS MÉDICAS GENERADO EL: ' + hoy.strftime('%d/%m/%Y')

        ws.merge_cells('B1:E1')
        ws['B3'] = 'RUN'
        ws['C3'] = 'FUNCIONARIO'
        ws['D3'] = 'INICIO'
        ws['E3'] = 'FIN'
        ws['F3'] = 'DIAS'

        cont = 4

        for ausencia in ausencias:
            if ausencia.tipo == 'LM':
                nombre = ausencia.funcionario.nombres +\
                         ' ' + ausencia.funcionario.apellido_paterno +\
                         ' ' + ausencia.funcionario.apellido_materno
                ws.cell(row=cont, column=2).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=3).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=4).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=5).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=6).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=2).value = ausencia.funcionario.run
                ws.cell(row=cont, column=3).value = nombre.upper()
                if ausencia.fecha_inicio is not None:
                    ws.cell(row=cont, column=4).value = ausencia.fecha_inicio.strftime('%d/%m/%Y')
                else:
                    ws.cell(row=cont, column=4).value = 'No Ingresado'

                if ausencia.fecha_termino is not None:
                    ws.cell(row=cont, column=5).value = ausencia.fecha_termino.strftime('%d/%m/%Y')
                else:
                    ws.cell(row=cont, column=5).value = 'No Ingresado'

                ws.cell(row=cont, column=6).value = ausencia.dias
                cont += 1

        nombre_archivo = "Reporte_Licencias_Mes_Actual_{}.xlsx".format(hoy)
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


# REPORTE EXCEL DE LICENCIAS DEL MES ANTERIOR
@method_decorator(login_required(), name='dispatch')
class ReporteLicenciasExcelMesPasado(TemplateView):
    def get(self, request, *args, **kwargs):
        hoy = datetime.date.today()
        annio = hoy.year
        mes = hoy.month
        mes = hoy+relativedelta(months=-1)
        ant_mes = mes.month
        ausencias = Ausencia.objects.filter(fecha_inicio__year=annio).filter(fecha_inicio__month=ant_mes)
        wb = Workbook()
        ws = wb.active
        b1 = ws['B1']
        b3 = ws['B3']
        c3 = ws['C3']
        d3 = ws['D3']
        e3 = ws['E3']
        f3 = ws['F3']
        # TRABAJANDO EN EL REPORTE EXCEL
        ftr = Font(color=colors.BLUE, bold=True)
        ftb = Font(color=colors.BLACK, bold=True, size=10)
        b1.font = ftr
        b3.font = ftb
        c3.font = ftb
        d3.font = ftb
        e3.font = ftb

        ws['B1'] = 'REPORTE DE LICENCIAS GENERADO EL : ' + hoy.strftime('%d/%m/%Y')

        ws.merge_cells('B1:E1')
        ws['B3'] = 'RUN'
        ws['C3'] = 'FUNCIONARIO'
        ws['D3'] = 'INICIO'
        ws['E3'] = 'FIN'

        cont = 4

        for ausencia in ausencias:
            if ausencia.tipo == 'LM':
                nombre = ausencia.funcionario.nombres +\
                         ' ' + ausencia.funcionario.apellido_paterno +\
                         ' ' + ausencia.funcionario.apellido_materno
                ws.cell(row=cont, column=2).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=3).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=4).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=5).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=6).font = Font(color=colors.BLACK, size=10)
                ws.cell(row=cont, column=2).value = ausencia.funcionario.run
                ws.cell(row=cont, column=3).value = nombre.upper()
                if ausencia.fecha_inicio is not None:
                    ws.cell(row=cont, column=4).value = ausencia.fecha_inicio.strftime('%d/%m/%Y')
                else:
                    ws.cell(row=cont, column=4).value = 'No Ingresado'
                if ausencia.fecha_termino is not None:
                    ws.cell(row=cont, column=5).value = ausencia.fecha_termino.strftime('%d/%m/%Y')
                else:
                    ws.cell(row=cont, column=5).value = 'No Ingresado'
                ws.cell(row=cont, column=6).value = ausencia.dias
                cont += 1

        nombre_archivo = "Reporte_Licencias_Mes_Anterior_{}.xlsx".format(hoy)
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


# REPORTES DE INDUCCIONES ########################################################


# REPORTE EN EXCEL DE TODAS LAS INDUCCIONES DESDE EL 2016
@method_decorator(login_required(), name='dispatch')
class ReporteInduccionesExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        hoy = datetime.date.today()
        mes = mes = hoy.month
        annio = hoy.year
        funcionarios = Funcionario.objects.filter(fecha_induccion__year=2016)
        wb = Workbook()
        ws = wb.active
        # TRABAJANDO EN EL REPORTE EXCEL
        b1 = ws['B1']
        b3 = ws['B3']
        c3 = ws['C3']
        d3 = ws['D3']
        e3 = ws['E3']

        ftr = Font(color=colors.BLUE, bold=True, size=10)
        ftb = Font(color=colors.BLACK, bold=True, size=10)
        b1.font = ftr
        b3.font = ftb
        c3.font = ftb
        d3.font = ftb
        e3.font = ftb

        ws['B1'] = 'REPORTE DE INDUCCIONES GENERADO EL: ' + hoy.strftime('%d/%m/%Y')

        ws.merge_cells('B1:E1')
        ws['B3'] = 'RUN'
        ws['C3'] = 'FUNCIONARIO'
        ws['D3'] = 'INGRESO'
        ws['E3'] = 'INDUCCIÓN'

        cont = 4

        for funcionario in funcionarios:
            nombre = funcionario.nombres +\
                     ' ' + funcionario.apellido_paterno +\
                     ' ' + funcionario.apellido_materno
            ws.cell(row=cont, column=2).font = Font(color=colors.BLACK, size=10)
            ws.cell(row=cont, column=3).font = Font(color=colors.BLACK, size=10)
            ws.cell(row=cont, column=4).font = Font(color=colors.BLACK, size=10)
            ws.cell(row=cont, column=5).font = Font(color=colors.BLACK, size=10)
            ws.cell(row=cont, column=2).value = funcionario.run
            ws.cell(row=cont, column=3).value = nombre.upper()
            if funcionario.fecha_ingreso is not None:
                ws.cell(row=cont, column=4).value = funcionario.fecha_ingreso.strftime('%d/%m/%Y')
            else:
                ws.cell(row=cont, column=4).value = 'No Ingresado'
            if funcionario.fecha_induccion is not None:
                ws.cell(row=cont, column=5).value = funcionario.fecha_induccion.strftime('%d/%m/%Y')
            else:
                ws.cell(row=cont, column=5).value = 'No Ingresado'
            cont += 1

        nombre_archivo = "Reporte_Inducciones_Completo_{}.xlsx".format(hoy)
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


# REPORTE DEL TOTAL DE TRABAJADORES DEL CESFAM
@method_decorator(login_required(), name='dispatch')
class ReporteTotalExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        hoy = datetime.date.today()
        funcionarios = Funcionario.objects.all()
        wb = Workbook()
        ws = wb.active
        # TRABAJANDO EN EL REPORTE EXCEL
        b1 = ws['B1']
        b3 = ws['B3']
        c3 = ws['C3']
        d3 = ws['D3']
        e3 = ws['E3']

        ftr = Font(color=colors.BLUE, bold=True, size=10)
        ftb = Font(color=colors.BLACK, bold=True, size=10)
        b1.font = ftr
        b3.font = ftb
        c3.font = ftb
        d3.font = ftb
        e3.font = ftb

        ws['B1'] = 'REPORTE DE INDUCCIONES GENERADO EL: ' + hoy.strftime('%d/%m/%Y')

        ws.merge_cells('B1:E1')
        ws['B3'] = 'RUN'
        ws['C3'] = 'FUNCIONARIO'
        ws['D3'] = 'INGRESO'
        ws['E3'] = 'INDUCCIÓN'

        cont = 4

        for funcionario in funcionarios:
            nombre = funcionario.nombres + \
                     ' ' + funcionario.apellido_paterno + \
                     ' ' + funcionario.apellido_materno
            ws.cell(row=cont, column=2).font = Font(color=colors.BLACK, size=10)
            ws.cell(row=cont, column=3).font = Font(color=colors.BLACK, size=10)
            ws.cell(row=cont, column=4).font = Font(color=colors.BLACK, size=10)
            ws.cell(row=cont, column=5).font = Font(color=colors.BLACK, size=10)
            ws.cell(row=cont, column=2).value = funcionario.run
            ws.cell(row=cont, column=3).value = nombre.upper()
            if funcionario.fecha_ingreso is not None:
                ws.cell(row=cont, column=4).value = funcionario.fecha_ingreso.strftime('%d/%m/%Y')
            else:
                ws.cell(row=cont, column=4).value = 'No Indica'
            if funcionario.fecha_induccion is not None:
                ws.cell(row=cont, column=5).value = funcionario.fecha_induccion.strftime('%d/%m/%Y')
            else:
                ws.cell(row=cont, column=5).value = 'No Indica'
            cont += 1

        nombre_archivo = "Reporte_Total_{}.xlsx".format(hoy)
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
